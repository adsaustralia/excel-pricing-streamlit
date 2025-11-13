
import streamlit as st
import openpyxl
import pandas as pd
import re
from io import BytesIO
from openpyxl.utils import column_index_from_string, get_column_letter

st.set_page_config(page_title="Excel SQM & Pricing Tool", layout="wide")
st.title("📊 Excel SQM & Pricing Calculator — Multi-Sheet Version")

st.write('''
Upload your Excel file, define column/row settings, and input pricing rules.  
The app will calculate **SQM and prices** for each sheet, display previews, and let you **download results or a combined summary**.
''')

# -----------------------------
# Helper Functions
# -----------------------------
def normalize(s):
    return re.sub(r'[^a-z0-9]+', '', str(s).lower()) if s else ""

def get_rate(material, prices):
    m = normalize(material)
    if "jellyfish" in m:
        return prices["Jellyfish"]
    if "ferrous" in m:
        return prices["Ferrous"]
    if "syntheticbanner" in m and "ds" in m:
        return prices["Synthetic Banner (DS)"]
    if "syntheticbanner" in m and "ss" in m:
        return prices["Synthetic Banner (SS)"]
    if "backlitshimmer" in m:
        return prices["Backlit Shimmer"]
    if "artboard" in m or "bflute" in m or "400gsm" in m:
        return prices["Artboard"]
    return None

def parse_size(raw):
    if not raw:
        return None, None
    s = str(raw).replace("×", "x").replace("X", "x").replace("*", "x")
    nums = re.findall(r'\d+(?:\.\d+)?', s)
    if len(nums) >= 2:
        return float(nums[0]), float(nums[1])
    return None, None

def parse_qty(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    m = re.search(r'\d+(?:\.\d+)?', str(raw).replace(",", ""))
    return float(m.group(0)) if m else None

def clean_value(v):
    '''
    Clean raw cell value:
    - If it's a formula string like ='PRINT DB'!$G13 or =SUM(A1:A5), ignore it.
    - Otherwise return as-is.
    '''
    if v is None:
        return None
    if isinstance(v, str) and v.strip().startswith("="):
        return None
    return v

# -----------------------------
# File Upload
# -----------------------------
uploaded_file = st.file_uploader("📤 Upload Excel file", type=["xlsx"])

if uploaded_file:
    # data_only=True -> use cached formula results instead of formula text
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet_names = wb.sheetnames

    # -----------------------------
    # Configuration
    # -----------------------------
    st.subheader("⚙️ Excel Structure Settings")

    col1, col2 = st.columns(2)
    with col1:
        start_col = st.text_input("Start Column", value="AC")
        end_col = st.text_input("End Column", value="IG")
    with col2:
        row_size = st.number_input("Row (Size)", value=5)
        row_material = st.number_input("Row (Material)", value=6)
        row_qty = st.number_input("Row (Quantity)", value=155)
        row_sqm = st.number_input("Row (SQM Output)", value=156)
        row_price = st.number_input("Row (Price Output)", value=157)

    # -----------------------------
    # Pricing Inputs
    # -----------------------------
    st.subheader("💰 Material Pricing Rules (AUD/m²)")

    col1, col2, col3 = st.columns(3)
    with col1:
        price_jellyfish = st.number_input("Jellyfish", value=7.55)
        price_ferrous = st.number_input("Ferrous", value=12.00)
    with col2:
        price_artboard = st.number_input("Artboard / B-Flute / 400gsm", value=6.42)
        price_ss = st.number_input("Synthetic Banner (SS)", value=13.00)
    with col3:
        price_ds = st.number_input("Synthetic Banner (DS)", value=19.50)
        price_backlit = st.number_input("Backlit Shimmer", value=40.00)

    prices = {
        "Jellyfish": price_jellyfish,
        "Ferrous": price_ferrous,
        "Artboard": price_artboard,
        "Synthetic Banner (SS)": price_ss,
        "Synthetic Banner (DS)": price_ds,
        "Backlit Shimmer": price_backlit
    }

    # -----------------------------
    # Single or All Sheets Option
    # -----------------------------
    st.subheader("📄 Sheet Selection")
    process_all = st.checkbox("Process all sheets", value=True)

    sheet_choice = None
    if not process_all:
        sheet_choice = st.selectbox("Select a sheet", sheet_names)

    # -----------------------------
    # Process Button
    # -----------------------------
    if st.button("🚀 Process & Calculate"):
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        summary_data = []

        # Function to process a single sheet
        def process_sheet(ws, sheet_name):
            total_price = 0
            preview_data = []

            for c in range(start_idx, end_idx + 1):
                col = get_column_letter(c)

                # Read and clean raw values from Excel
                raw_size = ws[f"{col}{row_size}"].value
                raw_material = ws[f"{col}{row_material}"].value
                raw_qty = ws[f"{col}{row_qty}"].value

                size_val = clean_value(raw_size)
                material_val = clean_value(raw_material)
                qty_val = clean_value(raw_qty)

                w, h = parse_size(size_val)
                qty = parse_qty(qty_val)
                rate = get_rate(material_val, prices)
                sqm, price = None, None

                if w and h and qty:
                    sqm = (w / 1000) * (h / 1000) * qty
                    ws[f"{col}{row_sqm}"].value = round(sqm, 6)
                    if rate:
                        price = round(sqm * rate, 2)
                        ws[f"{col}{row_price}"].value = price
                        total_price += price

                preview_data.append({
                    "Column": col,
                    "Material": material_val,
                    "Size": size_val,
                    "Qty": qty,
                    "Rate": rate,
                    "SQM": round(sqm, 3) if sqm else None,
                    "Price (AUD)": price
                })

            # Write sheet total & labels
            ws[f"{end_col}{row_sqm}"] = "TOTAL"
            ws[f"{end_col}{row_price}"] = round(total_price, 2)
            ws["AB" + str(row_sqm)] = "SQM"
            ws["AB" + str(row_price)] = "PRICE (AUD)"

            summary_data.append({"Sheet": sheet_name, "Total (AUD)": round(total_price, 2)})
            return pd.DataFrame(preview_data), total_price

        # Process all sheets or one
        if process_all:
            for name in sheet_names:
                ws_sheet = wb[name]
                df, total = process_sheet(ws_sheet, name)
                st.markdown(f"### 🧾 Preview — {name}")
                st.dataframe(df.head(15))
                st.info(f"Subtotal for {name}: ${total:,.2f}")
        else:
            ws_sheet = wb[sheet_choice]
            df, total = process_sheet(ws_sheet, sheet_choice)
            st.markdown(f"### 🧾 Preview — {sheet_choice}")
            st.dataframe(df.head(15))
            st.info(f"Total for {sheet_choice}: ${total:,.2f}")

        # Combined Summary
        df_summary = pd.DataFrame(summary_data)
        st.subheader("📘 Combined Totals Summary")
        st.dataframe(df_summary)

        grand_total = df_summary["Total (AUD)"].sum()
        st.success(f"✅ Grand Total across all sheets: ${grand_total:,.2f}")

        # Save updated workbook
        output_excel = BytesIO()
        wb.save(output_excel)
        output_excel.seek(0)

        # Save summary table separately
        output_summary = BytesIO()
        with pd.ExcelWriter(output_summary, engine='openpyxl') as writer:
            df_summary.to_excel(writer, index=False, sheet_name="Summary")
        output_summary.seek(0)

        st.download_button(
            label="⬇️ Download Updated Excel Workbook",
            data=output_excel,
            file_name="Updated_Pricing_All_Sheets.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.download_button(
            label="⬇️ Download Summary Totals Only",
            data=output_summary,
            file_name="Pricing_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
